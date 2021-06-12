import openpyxl
wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only = True)
#даю переменной значение документа ексель, дата_онли нужна чтобы читать значения ячеек, а не формулы в них

print(wb.sheetnames) #показую все листы

wb.active = 0 #выбераю активный лист

ws1 = wb.active #даю переменной значение активного листа

for i in range(1,11):
    a = ws1['B'+str(i)].value #даю переменной значение ячейки
    print(a)
####################################

ws2 = wb.create_sheet(title="Try")

####################################
wb.save("sample.xlsx")