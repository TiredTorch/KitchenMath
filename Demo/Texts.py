import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.reader.excel.load_workbook(filename="sample.xlsx", data_only = True)

wb.active = 0
ws = wb.active

def Test():

    #ws['F2'].value = "ff\nff"
    #ws['F2'].value += "aaaa"

    #ws['F2'].alignment = Alignment(text_rotation=90, horizontal='center')


    if(ws['F2'].value):
        print(1)
    else:
        print(0)

    wb.save('sample.xlsx')

Test()

def PrepareTable(letters, wsT):
    wsT['D17'].value = str('')
    wsT['D11'].value = str('')
    wsT['H17'].value = str('')
    wsT['H11'].value = str('')
    wsT['L17'].value = str('')
    wsT['L11'].value = str('')
    for letter in letters:
        for num in range (19, 81):
            wsT[letter+str(num)].value = float()

    wsT['S84'].value = '=SUM(R19:S81)'

    for i in range(19, 81):
        wsT['P'+str(i)].value = '=SUM(D'+str(i)+':O'+str(i)+')/1000'
        wsT['V'+str(i)].value = '=P'+str(i)+'*B5'
